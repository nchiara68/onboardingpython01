"""
🏢 Working Factoring Analysis Excel Export
Save this file as: factoring_analysis/scripts/working_export.py
"""

import pandas as pd
import numpy as np
import sys
import os
from datetime import datetime

# Add the scripts directory to Python path
sys.path.append(os.path.dirname(__file__))

try:
    from factoring_analyzer import FactoringAnalyzer
except ImportError:
    print("❌ Could not import FactoringAnalyzer. Make sure factoring_analyzer.py is in the scripts directory.")
    sys.exit(1)

def sort_aging_buckets(df, aging_col='Aging_Bucket'):
    """Sort aging buckets in logical order"""
    aging_order = ['On-time', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', '90+ Days']
    
    if aging_col in df.columns:
        df[aging_col] = pd.Categorical(df[aging_col], categories=aging_order, ordered=True)
        df = df.sort_values(aging_col)
        df[aging_col] = df[aging_col].astype(str)
    
    return df

def working_export():
    """Working export function that definitely creates Excel file"""
    print("🏢 WORKING FACTORING ANALYSIS EXPORT")
    print("="*60)
    
    # File paths
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(current_dir)
    csv_path = os.path.join(project_root, 'data', 'TecnoCargoInvoiceDataset01.csv')
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"factoring_analysis_working_{timestamp}.xlsx"
    output_path = os.path.join(project_root, 'output', output_filename)
    
    # Create output directory
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    print(f"📁 Input: {csv_path}")
    print(f"📁 Output: {output_path}")
    
    # Load analyzer
    print("\n📊 Loading data...")
    analyzer = FactoringAnalyzer(csv_path)
    
    # Create data separations MANUALLY to ensure they exist
    print("🔄 Creating data separations...")
    
    # STEP 1: Create paid vs outstanding invoices
    df_paid_invoices = analyzer.df_invoice[analyzer.df_invoice['Amt. Due (USD)'] == 0].copy()
    df_outstanding_invoices = analyzer.df_invoice[analyzer.df_invoice['Amt. Due (USD)'] > 0].copy()
    
    print(f"   📈 Paid invoices: {len(df_paid_invoices)}")
    print(f"   ⏳ Outstanding invoices: {len(df_outstanding_invoices)}")
    print(f"   💳 Credit memos: {len(analyzer.df_credit_memo)}")
    
    # STEP 2: Ensure aging buckets exist for both
    for df_name, df in [("Paid", df_paid_invoices), ("Outstanding", df_outstanding_invoices)]:
        if len(df) > 0 and 'Aging_Bucket' not in df.columns:
            print(f"   🔧 Adding aging buckets to {df_name} invoices...")
            df['Days_Past_Due'] = (datetime.now() - df['Due Date']).dt.days
            df['Aging_Bucket'] = pd.cut(
                df['Days_Past_Due'],
                bins=[-float('inf'), 0, 30, 60, 90, float('inf')],
                labels=['Current', '1-30 Days', '31-60 Days', '61-90 Days', '90+ Days']
            )
    
    # STEP 3: Create Excel file
    print("\n📊 Creating Excel file...")
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # ================================================================
        # ALWAYS CREATE SUMMARY SHEET FIRST
        # ================================================================
        print("   📋 Creating Summary sheet...")
        summary_data = {
            'Category': ['Total Records', 'Invoices', 'Paid Invoices', 'Outstanding Invoices', 'Credit Memos'],
            'Count': [
                len(analyzer.df_full),
                len(analyzer.df_invoice),
                len(df_paid_invoices),
                len(df_outstanding_invoices),
                len(analyzer.df_credit_memo)
            ],
            'Total Amount': [
                analyzer.df_full['Amount (USD)'].sum(),
                analyzer.df_invoice['Amount (USD)'].sum(),
                df_paid_invoices['Amount (USD)'].sum() if len(df_paid_invoices) > 0 else 0,
                df_outstanding_invoices['Amt. Due (USD)'].sum() if len(df_outstanding_invoices) > 0 else 0,
                analyzer.df_credit_memo['Amount (USD)'].sum() if len(analyzer.df_credit_memo) > 0 else 0
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        print("   ✅ Summary sheet created")
        
        # ================================================================
        # PAID INVOICES ANALYSIS
        # ================================================================
        if len(df_paid_invoices) > 0:
            print("   📈 Creating Paid Invoice tables...")
            
            # Table 1.1: Count by Aging
            paid_count = df_paid_invoices.groupby('Aging_Bucket')['Number'].count().reset_index()
            paid_count.columns = ['Aging Bucket', 'Invoice Count']
            paid_count = sort_aging_buckets(paid_count)
            paid_count.to_excel(writer, sheet_name='1.1_Paid_Count_by_Aging', index=False)
            
            # Table 1.2: Amount by Aging
            paid_amount = df_paid_invoices.groupby('Aging_Bucket')['Amount (USD)'].sum().reset_index()
            paid_amount.columns = ['Aging Bucket', 'Total Amount']
            paid_amount = sort_aging_buckets(paid_amount)
            paid_amount.to_excel(writer, sheet_name='1.2_Paid_Amount_by_Aging', index=False)
            
            # Table 1.3: Percentages
            total_paid_count = len(df_paid_invoices)
            total_paid_amount = df_paid_invoices['Amount (USD)'].sum()
            
            paid_pct = df_paid_invoices.groupby('Aging_Bucket').agg({
                'Number': 'count',
                'Amount (USD)': 'sum'
            }).reset_index()
            
            paid_pct['Count Percentage'] = (paid_pct['Number'] / total_paid_count * 100).round(2)
            paid_pct['Amount Percentage'] = (paid_pct['Amount (USD)'] / total_paid_amount * 100).round(2)
            paid_pct_final = paid_pct[['Aging_Bucket', 'Count Percentage', 'Amount Percentage']].copy()
            paid_pct_final.columns = ['Aging Bucket', 'Count Percentage', 'Amount Percentage']
            paid_pct_final = sort_aging_buckets(paid_pct_final)
            paid_pct_final.to_excel(writer, sheet_name='1.3_Paid_Percentages', index=False)
            
            # Table 1.4: Customer Analysis (Simple version first)
            try:
                customer_paid = df_paid_invoices.groupby('Applied to').agg({
                    'Number': 'count',
                    'Amount (USD)': 'sum'
                }).reset_index()
                customer_paid.columns = ['Customer', 'Invoice Count', 'Total Amount']
                customer_paid = customer_paid.sort_values('Total Amount', ascending=False).head(20)
                customer_paid.to_excel(writer, sheet_name='1.4_Paid_Top_Customers', index=False)
                
                # Also create customer aging pivot if possible
                try:
                    customer_aging_pivot = pd.crosstab(df_paid_invoices['Applied to'], 
                                                     df_paid_invoices['Aging_Bucket'], 
                                                     values=df_paid_invoices['Amount (USD)'], 
                                                     aggfunc='sum').fillna(0)
                    customer_aging_pivot.to_excel(writer, sheet_name='1.4_Paid_Customer_Aging_Matrix')
                except:
                    pass
                    
            except Exception as e:
                print(f"   ⚠️  Customer analysis skipped: {e}")
            
            print("   ✅ Paid invoice tables created")
        
        # ================================================================
        # OUTSTANDING INVOICES ANALYSIS
        # ================================================================
        if len(df_outstanding_invoices) > 0:
            print("   ⏳ Creating Outstanding Invoice tables...")
            
            # Table 2.1: Count by Aging
            out_count = df_outstanding_invoices.groupby('Aging_Bucket')['Number'].count().reset_index()
            out_count.columns = ['Aging Bucket', 'Invoice Count']
            out_count = sort_aging_buckets(out_count)
            out_count.to_excel(writer, sheet_name='2.1_Outstanding_Count_Aging', index=False)
            
            # Table 2.2: Amount by Aging
            out_amount = df_outstanding_invoices.groupby('Aging_Bucket')['Amt. Due (USD)'].sum().reset_index()
            out_amount.columns = ['Aging Bucket', 'Total Amount Due']
            out_amount = sort_aging_buckets(out_amount)
            out_amount.to_excel(writer, sheet_name='2.2_Outstanding_Amount_Aging', index=False)
            
            # Table 2.3: Percentages
            total_out_count = len(df_outstanding_invoices)
            total_out_amount = df_outstanding_invoices['Amt. Due (USD)'].sum()
            
            out_pct = df_outstanding_invoices.groupby('Aging_Bucket').agg({
                'Number': 'count',
                'Amt. Due (USD)': 'sum'
            }).reset_index()
            
            out_pct['Count Percentage'] = (out_pct['Number'] / total_out_count * 100).round(2)
            out_pct['Amount Percentage'] = (out_pct['Amt. Due (USD)'] / total_out_amount * 100).round(2)
            out_pct_final = out_pct[['Aging_Bucket', 'Count Percentage', 'Amount Percentage']].copy()
            out_pct_final.columns = ['Aging Bucket', 'Count Percentage', 'Amount Percentage']
            out_pct_final = sort_aging_buckets(out_pct_final)
            out_pct_final.to_excel(writer, sheet_name='2.3_Outstanding_Percentages', index=False)
            
            # Table 2.4: Customer Analysis
            try:
                customer_outstanding = df_outstanding_invoices.groupby('Applied to').agg({
                    'Number': 'count',
                    'Amt. Due (USD)': 'sum'
                }).reset_index()
                customer_outstanding.columns = ['Customer', 'Invoice Count', 'Total Amount Due']
                customer_outstanding = customer_outstanding.sort_values('Total Amount Due', ascending=False).head(20)
                customer_outstanding.to_excel(writer, sheet_name='2.4_Outstanding_Top_Customers', index=False)
                
                # Also create customer aging pivot if possible
                try:
                    customer_out_pivot = pd.crosstab(df_outstanding_invoices['Applied to'], 
                                                   df_outstanding_invoices['Aging_Bucket'], 
                                                   values=df_outstanding_invoices['Amt. Due (USD)'], 
                                                   aggfunc='sum').fillna(0)
                    customer_out_pivot.to_excel(writer, sheet_name='2.4_Outstanding_Customer_Matrix')
                except:
                    pass
                    
            except Exception as e:
                print(f"   ⚠️  Customer analysis skipped: {e}")
            
            print("   ✅ Outstanding invoice tables created")
        
        # ================================================================
        # CREDIT MEMO ANALYSIS
        # ================================================================
        if len(analyzer.df_credit_memo) > 0:
            print("   💳 Creating Credit Memo tables...")
            
            # Table 3.1: Credit by Customer
            credit_by_customer = analyzer.df_credit_memo.groupby('Applied to')['Amount (USD)'].sum().reset_index()
            credit_by_customer.columns = ['Customer', 'Total Credit Amount']
            credit_by_customer = credit_by_customer.sort_values('Total Credit Amount', ascending=False)
            
            # Add percentage
            total_credit_amount = credit_by_customer['Total Credit Amount'].sum()
            credit_by_customer['Percentage of Total'] = (credit_by_customer['Total Credit Amount'] / total_credit_amount * 100).round(2)
            credit_by_customer.to_excel(writer, sheet_name='3.1_Credit_by_Customer', index=False)
            
            # Table 3.2: Credit Summary
            credit_stats = pd.DataFrame({
                'Metric': ['Total Credit Memos', 'Total Credit Amount', 'Average Credit Amount', 
                          'Largest Credit Memo', 'Number of Customers with Credits'],
                'Value': [
                    len(analyzer.df_credit_memo),
                    analyzer.df_credit_memo['Amount (USD)'].sum(),
                    analyzer.df_credit_memo['Amount (USD)'].mean(),
                    analyzer.df_credit_memo['Amount (USD)'].max(),
                    analyzer.df_credit_memo['Applied to'].nunique()
                ]
            })
            credit_stats.to_excel(writer, sheet_name='3.2_Credit_Summary', index=False)
            
            print("   ✅ Credit memo tables created")
        
        # ================================================================
        # RAW DATA SHEETS
        # ================================================================
        print("   📋 Adding raw data sheets...")
        
        # Always add invoice data
        analyzer.df_invoice.to_excel(writer, sheet_name='RAW_All_Invoices', index=False)
        
        # Add separated data if exists
        if len(df_paid_invoices) > 0:
            df_paid_invoices.to_excel(writer, sheet_name='RAW_Paid_Invoices', index=False)
        
        if len(df_outstanding_invoices) > 0:
            df_outstanding_invoices.to_excel(writer, sheet_name='RAW_Outstanding_Invoices', index=False)
        
        if len(analyzer.df_credit_memo) > 0:
            analyzer.df_credit_memo.to_excel(writer, sheet_name='RAW_Credit_Memos', index=False)
        
        print("   ✅ Raw data sheets added")
    
    print(f"\n🎉 Excel file created successfully!")
    print(f"📁 File location: {output_path}")
    
    # Count and report sheets
    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    sheet_count = len(wb.sheetnames)
    print(f"📊 Total worksheets created: {sheet_count}")
    print(f"📋 Sheet names: {wb.sheetnames}")
    
    return output_path

if __name__ == "__main__":
    try:
        output_file = working_export()
        print("\n✅ SUCCESS! Export completed without errors.")
    except Exception as e:
        print(f"\n❌ ERROR: Export failed with error: {e}")
        import traceback
        traceback.print_exc()
        