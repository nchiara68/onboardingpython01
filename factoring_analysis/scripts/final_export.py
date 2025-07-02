"""
🏢 Final Working Factoring Analysis Excel Export
Save this file as: factoring_analysis/scripts/final_export.py
"""

import pandas as pd
import numpy as np
import sys
import os
from datetime import datetime

# Add the scripts directory to Python path
sys.path.append(os.path.dirname(__file__))

try:
    from factoring_analyzer import FactoringAnalyzer
except ImportError:
    print("❌ Could not import FactoringAnalyzer. Make sure factoring_analyzer.py is in the scripts directory.")
    sys.exit(1)

def sort_aging_buckets(df, aging_col='Aging_Bucket'):
    """Sort aging buckets in logical order"""
    aging_order = ['On-time', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', '90+ Days']
    
    if aging_col in df.columns:
        df[aging_col] = pd.Categorical(df[aging_col], categories=aging_order, ordered=True)
        df = df.sort_values(aging_col)
        df[aging_col] = df[aging_col].astype(str)
    
    return df

def calculate_aging_correctly(df, invoice_type):
    """Calculate aging delay days correctly based on invoice type"""
    if len(df) == 0:
        return df
    
    print(f"      🔧 Calculating aging delay days for {invoice_type} invoices...")
    
    # Ensure date columns are datetime
    df['Due Date'] = pd.to_datetime(df['Due Date'], errors='coerce')
    if 'Last Payment Date' in df.columns:
        df['Last Payment Date'] = pd.to_datetime(df['Last Payment Date'], errors='coerce')
    
    # Set cutoff date - June 23, 2025
    cutoff_date = pd.Timestamp('2025-06-23')
    
    if invoice_type == 'paid':
        # For paid invoices: Aging Delay Days = Last Payment Date - Due Date
        print(f"         Using Last Payment Date - Due Date for paid invoices...")
        df['Aging Delay Days'] = (df['Last Payment Date'] - df['Due Date']).dt.days
        # Handle missing Last Payment Date
        df['Aging Delay Days'] = df['Aging Delay Days'].fillna(0)
    
    elif invoice_type == 'outstanding':
        # For outstanding invoices: Aging Delay Days = June 23, 2025 - Due Date
        print(f"         Using June 23, 2025 - Due Date for outstanding invoices...")
        df['Aging Delay Days'] = (cutoff_date - df['Due Date']).dt.days
        # Handle missing Due Date
        df['Aging Delay Days'] = df['Aging Delay Days'].fillna(0)
    
    # Create aging buckets based on aging delay days
    def create_aging_bucket(days, invoice_type):
        if days <= 0:
            return 'On-time' if invoice_type == 'paid' else 'Current'
        elif 1 <= days <= 30:
            return '1-30 Days'
        elif 31 <= days <= 60:
            return '31-60 Days'
        elif 61 <= days <= 90:
            return '61-90 Days'
        else:
            return '90+ Days'
    
    print(f"         Creating aging buckets...")
    df['Aging_Bucket'] = df['Aging Delay Days'].apply(lambda x: create_aging_bucket(x, invoice_type))
    
    # Clean up the dataframe - keep only original columns + aging columns
    # Original columns up to Amt. Due (USD), then add aging columns
    original_cols = ['Type', 'Number', 'Transaction Date', 'Applied to', 'Amount (USD)', 
                    'Due Date', 'Status', 'Last Payment Date', 'Amt. Paid (USD)', 'Amt. Due (USD)']
    
    # Keep only columns that exist in the dataframe
    cols_to_keep = [col for col in original_cols if col in df.columns]
    cols_to_keep.extend(['Aging Delay Days', 'Aging_Bucket'])
    
    df_clean = df[cols_to_keep].copy()
    
    print(f"         ✅ {invoice_type} aging calculation completed")
    print(f"         📊 Aging bucket distribution: {df_clean['Aging_Bucket'].value_counts().to_dict()}")
    
    return df_clean

def final_export():
    """Final working export function"""
    print("🏢 FINAL FACTORING ANALYSIS EXPORT")
    print("="*60)
    
    # File paths
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(current_dir)
    csv_path = os.path.join(project_root, 'data', 'TecnoCargoInvoiceDataset01.csv')
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"factoring_analysis_final_{timestamp}.xlsx"
    output_path = os.path.join(project_root, 'output', output_filename)
    
    # Create output directory
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    print(f"📁 Input: {csv_path}")
    print(f"📁 Output: {output_path}")
    
    # Load analyzer
    print("\n📊 Loading data with FactoringAnalyzer...")
    analyzer = FactoringAnalyzer(csv_path)
    
    # Use the original invoice dataset (before aging calculations)
    base_df = analyzer.df_invoice.copy()  # This is the clean invoice dataset
    
    print(f"📋 Base dataset shape: {base_df.shape}")
    print(f"📋 Columns available: {list(base_df.columns)}")
    
    # Create data separations using the base dataset
    print("\n🔄 Creating data separations and calculating aging correctly...")
    
    # STEP 1: Create paid vs outstanding invoices from the base dataset
    df_paid_invoices_raw = base_df[base_df['Amt. Due (USD)'] == 0].copy()
    df_outstanding_invoices_raw = base_df[base_df['Amt. Due (USD)'] > 0].copy()
    
    print(f"   📈 Raw paid invoices: {len(df_paid_invoices_raw)}")
    print(f"   ⏳ Raw outstanding invoices: {len(df_outstanding_invoices_raw)}")
    print(f"   💳 Credit memos: {len(analyzer.df_credit_memo)}")
    
    # STEP 2: Calculate aging correctly for each type
    df_paid_invoices = calculate_aging_correctly(df_paid_invoices_raw, 'paid')
    df_outstanding_invoices = calculate_aging_correctly(df_outstanding_invoices_raw, 'outstanding')
    
    print(f"\n📊 Final invoice counts after aging calculation:")
    print(f"   📈 Paid invoices: {len(df_paid_invoices)}")
    print(f"   ⏳ Outstanding invoices: {len(df_outstanding_invoices)}")
    
    # Verify aging buckets exist and show distribution
    if len(df_paid_invoices) > 0:
        print(f"   📊 Paid aging distribution: {df_paid_invoices['Aging_Bucket'].value_counts().to_dict()}")
    
    if len(df_outstanding_invoices) > 0:
        print(f"   📊 Outstanding aging distribution: {df_outstanding_invoices['Aging_Bucket'].value_counts().to_dict()}")
    
    # Create Excel file
    print("\n📊 Creating Excel file...")
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # ================================================================
            # SUMMARY SHEET - ALWAYS FIRST
            # ================================================================
            print("   📋 Creating Summary sheet...")
            summary_data = {
                'Metric': ['Total Records', 'Invoices', 'Paid Invoices', 'Outstanding Invoices', 'Credit Memos'],
                'Count': [
                    len(analyzer.df_full),
                    len(base_df),
                    len(df_paid_invoices),
                    len(df_outstanding_invoices),
                    len(analyzer.df_credit_memo)
                ],
                'Amount (USD)': [
                    analyzer.df_full['Amount (USD)'].sum(),
                    base_df['Amount (USD)'].sum(),
                    df_paid_invoices['Amount (USD)'].sum() if len(df_paid_invoices) > 0 else 0,
                    df_outstanding_invoices['Amt. Due (USD)'].sum() if len(df_outstanding_invoices) > 0 else 0,
                    analyzer.df_credit_memo['Amount (USD)'].sum() if len(analyzer.df_credit_memo) > 0 else 0
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            print("   ✅ Summary sheet created")
            
            # ================================================================
            # PAID INVOICES ANALYSIS
            # ================================================================
            if len(df_paid_invoices) > 0:
                print("   📈 Creating Paid Invoice analysis...")
                
                try:
                    # Table 1.1: Count by Aging
                    paid_count = df_paid_invoices.groupby('Aging_Bucket')['Number'].count().reset_index()
                    paid_count.columns = ['Aging_Bucket', 'Invoice Count']
                    paid_count = sort_aging_buckets(paid_count)
                    paid_count.to_excel(writer, sheet_name='1.1_Paid_Count_by_Aging', index=False)
                    
                    # Table 1.2: Amount by Aging
                    paid_amount = df_paid_invoices.groupby('Aging_Bucket')['Amount (USD)'].sum().reset_index()
                    paid_amount.columns = ['Aging_Bucket', 'Total Amount']
                    paid_amount = sort_aging_buckets(paid_amount)
                    paid_amount.to_excel(writer, sheet_name='1.2_Paid_Amount_by_Aging', index=False)
                    
                    # Table 1.3: Percentages
                    total_paid_count = len(df_paid_invoices)
                    total_paid_amount = df_paid_invoices['Amount (USD)'].sum()
                    
                    paid_pct = df_paid_invoices.groupby('Aging_Bucket').agg({
                        'Number': 'count',
                        'Amount (USD)': 'sum'
                    }).reset_index()
                    
                    paid_pct['Count Percentage'] = (paid_pct['Number'] / total_paid_count * 100).round(2)
                    paid_pct['Amount Percentage'] = (paid_pct['Amount (USD)'] / total_paid_amount * 100).round(2)
                    paid_pct_final = paid_pct[['Aging_Bucket', 'Count Percentage', 'Amount Percentage']].copy()
                    paid_pct_final.columns = ['Aging_Bucket', 'Count Percentage', 'Amount Percentage']
                    paid_pct_final = sort_aging_buckets(paid_pct_final)
                    paid_pct_final.to_excel(writer, sheet_name='1.3_Paid_Percentages', index=False)
                    
                    print("      ✅ Paid basic tables created")
                    
                    # Table 1.4: Customer Analysis
                    try:
                        # Simple customer summary
                        customer_paid = df_paid_invoices.groupby('Applied to').agg({
                            'Number': 'count',
                            'Amount (USD)': 'sum'
                        }).reset_index()
                        customer_paid.columns = ['Customer', 'Invoice Count', 'Total Amount']
                        customer_paid = customer_paid.sort_values('Total Amount', ascending=False).head(20)
                        customer_paid.to_excel(writer, sheet_name='1.4_Paid_Top_Customers', index=False)
                        
                        # Customer aging matrix
                        customer_aging_pivot = pd.crosstab(df_paid_invoices['Applied to'], 
                                                         df_paid_invoices['Aging_Bucket'], 
                                                         values=df_paid_invoices['Amount (USD)'], 
                                                         aggfunc='sum').fillna(0)
                        
                        # Get top 15 customers for the detailed matrix
                        top_customers = df_paid_invoices.groupby('Applied to')['Amount (USD)'].sum().nlargest(15).index
                        customer_aging_pivot_top = customer_aging_pivot.loc[top_customers]
                        customer_aging_pivot_top.to_excel(writer, sheet_name='1.4_Paid_Customer_Matrix')
                        
                        print("      ✅ Paid customer analysis created")
                        
                    except Exception as e:
                        print(f"      ⚠️  Paid customer analysis skipped: {e}")
                
                except Exception as e:
                    print(f"   ❌ Error in paid invoice analysis: {e}")
            
            # ================================================================
            # OUTSTANDING INVOICES ANALYSIS
            # ================================================================
            if len(df_outstanding_invoices) > 0:
                print("   ⏳ Creating Outstanding Invoice analysis...")
                
                try:
                    # Table 2.1: Count by Aging
                    out_count = df_outstanding_invoices.groupby('Aging_Bucket')['Number'].count().reset_index()
                    out_count.columns = ['Aging_Bucket', 'Invoice Count']
                    out_count = sort_aging_buckets(out_count)
                    out_count.to_excel(writer, sheet_name='2.1_Outstanding_Count_Aging', index=False)
                    
                    # Table 2.2: Amount by Aging
                    out_amount = df_outstanding_invoices.groupby('Aging_Bucket')['Amt. Due (USD)'].sum().reset_index()
                    out_amount.columns = ['Aging_Bucket', 'Total Amount Due']
                    out_amount = sort_aging_buckets(out_amount)
                    out_amount.to_excel(writer, sheet_name='2.2_Outstanding_Amount_Aging', index=False)
                    
                    # Table 2.3: Percentages
                    total_out_count = len(df_outstanding_invoices)
                    total_out_amount = df_outstanding_invoices['Amt. Due (USD)'].sum()
                    
                    out_pct = df_outstanding_invoices.groupby('Aging_Bucket').agg({
                        'Number': 'count',
                        'Amt. Due (USD)': 'sum'
                    }).reset_index()
                    
                    out_pct['Count Percentage'] = (out_pct['Number'] / total_out_count * 100).round(2)
                    out_pct['Amount Percentage'] = (out_pct['Amt. Due (USD)'] / total_out_amount * 100).round(2)
                    out_pct_final = out_pct[['Aging_Bucket', 'Count Percentage', 'Amount Percentage']].copy()
                    out_pct_final.columns = ['Aging_Bucket', 'Count Percentage', 'Amount Percentage']
                    out_pct_final = sort_aging_buckets(out_pct_final)
                    out_pct_final.to_excel(writer, sheet_name='2.3_Outstanding_Percentages', index=False)
                    
                    print("      ✅ Outstanding basic tables created")
                    
                    # Table 2.4: Customer Analysis
                    try:
                        # Simple customer summary
                        customer_outstanding = df_outstanding_invoices.groupby('Applied to').agg({
                            'Number': 'count',
                            'Amt. Due (USD)': 'sum'
                        }).reset_index()
                        customer_outstanding.columns = ['Customer', 'Invoice Count', 'Total Amount Due']
                        customer_outstanding = customer_outstanding.sort_values('Total Amount Due', ascending=False).head(20)
                        customer_outstanding.to_excel(writer, sheet_name='2.4_Outstanding_Top_Customers', index=False)
                        
                        # Customer aging matrix
                        customer_out_pivot = pd.crosstab(df_outstanding_invoices['Applied to'], 
                                                       df_outstanding_invoices['Aging_Bucket'], 
                                                       values=df_outstanding_invoices['Amt. Due (USD)'], 
                                                       aggfunc='sum').fillna(0)
                        
                        # Get top 15 customers for the detailed matrix
                        top_customers_out = df_outstanding_invoices.groupby('Applied to')['Amt. Due (USD)'].sum().nlargest(15).index
                        customer_out_pivot_top = customer_out_pivot.loc[top_customers_out]
                        customer_out_pivot_top.to_excel(writer, sheet_name='2.4_Outstanding_Customer_Matrix')
                        
                        print("      ✅ Outstanding customer analysis created")
                        
                    except Exception as e:
                        print(f"      ⚠️  Outstanding customer analysis skipped: {e}")
                
                except Exception as e:
                    print(f"   ❌ Error in outstanding invoice analysis: {e}")
            
            # ================================================================
            # CREDIT MEMO ANALYSIS
            # ================================================================
            if len(analyzer.df_credit_memo) > 0:
                print("   💳 Creating Credit Memo analysis...")
                
                try:
                    # Table 3.1: Credit by Customer
                    credit_by_customer = analyzer.df_credit_memo.groupby('Applied to')['Amount (USD)'].sum().reset_index()
                    credit_by_customer.columns = ['Customer', 'Total Credit Amount']
                    credit_by_customer = credit_by_customer.sort_values('Total Credit Amount', ascending=False)
                    
                    # Add percentage
                    total_credit_amount = credit_by_customer['Total Credit Amount'].sum()
                    credit_by_customer['Percentage of Total'] = (credit_by_customer['Total Credit Amount'] / total_credit_amount * 100).round(2)
                    credit_by_customer.to_excel(writer, sheet_name='3.1_Credit_by_Customer', index=False)
                    
                    # Table 3.2: Credit Summary
                    credit_stats = pd.DataFrame({
                        'Metric': ['Total Credit Memos', 'Total Credit Amount', 'Average Credit Amount', 
                                  'Largest Credit Memo', 'Number of Customers with Credits'],
                        'Value': [
                            len(analyzer.df_credit_memo),
                            analyzer.df_credit_memo['Amount (USD)'].sum(),
                            analyzer.df_credit_memo['Amount (USD)'].mean(),
                            analyzer.df_credit_memo['Amount (USD)'].max(),
                            analyzer.df_credit_memo['Applied to'].nunique()
                        ]
                    })
                    credit_stats.to_excel(writer, sheet_name='3.2_Credit_Summary', index=False)
                    
                    print("   ✅ Credit memo analysis created")
                
                except Exception as e:
                    print(f"   ❌ Error in credit memo analysis: {e}")
            
            # ================================================================
            # RAW DATA SHEETS
            # ================================================================
            print("   📋 Adding raw data sheets...")
            
            try:
                # Always add main invoice data with correct aging
                base_df.to_excel(writer, sheet_name='RAW_All_Invoices_Original', index=False)
                
                # Add separated data with correct aging calculations
                if len(df_paid_invoices) > 0:
                    df_paid_invoices.to_excel(writer, sheet_name='RAW_Paid_Invoices', index=False)
                
                if len(df_outstanding_invoices) > 0:
                    df_outstanding_invoices.to_excel(writer, sheet_name='RAW_Outstanding_Invoices', index=False)
                
                if len(analyzer.df_credit_memo) > 0:
                    analyzer.df_credit_memo.to_excel(writer, sheet_name='RAW_Credit_Memos', index=False)
                
                print("   ✅ Raw data sheets added")
                
            except Exception as e:
                print(f"   ❌ Error adding raw data: {e}")
        
        print(f"\n🎉 Excel file created successfully!")
        print(f"📁 File location: {output_path}")
        
        # Verify the file
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            sheet_count = len(wb.sheetnames)
            print(f"📊 Total worksheets created: {sheet_count}")
            print(f"📋 Sheet names: {wb.sheetnames[:10]}{'...' if len(wb.sheetnames) > 10 else ''}")
            wb.close()
        except:
            print("📊 File created but could not verify sheet count")
        
        return output_path
        
    except Exception as e:
        print(f"\n❌ ERROR during Excel creation: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    try:
        output_file = final_export()
        if output_file:
            print("\n✅ SUCCESS! Export completed without errors.")
            print(f"📁 Your Excel file is ready: {output_file}")
        else:
            print("\n❌ Export failed. Check error messages above.")
    except Exception as e:
        print(f"\n❌ CRITICAL ERROR: {e}")
        import traceback
        traceback.print_exc()